"""
Generate Excel compliance report v5 from new POC eval reports
(tmp/poc/1627_1777553113352_recrui8). Mirrors the 3-sheet layout
of custom_instructions_compliance_report_v4.xlsx.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Per-comment analysis hand-derived from each *_eval_report.md
COMMENT_ROWS = [
    # ApplicantDetailsPipeline.js
    ("FILE: ApplicantDetailsPipeline.js (server/utils/Jobs)", None, None, None, None, None, None, None, None, None),
    (
        "ApplicantDetailsPipeline.js", "C#1", 460, "Logic Bug", "major",
        "Affirms updated hasSalaryFilter correctly checks undefined/null before parsing, allowing salary=0 as a valid filter value.",
        "Ensure try/catch and undefined handling for API-related code",
        "Raised - Correctly (TP, Affirmative)",
        "Affirmative comment praising the PR's bug fix. Severity/category/line all correct (Grade A). No fix needed since the PR already contains the fix.",
        None,
    ),
    (
        "ApplicantDetailsPipeline.js", "C#2", 513, "Logic Bug", "major",
        "Affirms removal of parsedSalaryMin > 0 from $gte clause is a correct fix that allows zero as minimum salary.",
        "Ensure try/catch and undefined handling for API-related code",
        "Raised - Correctly (TP, Affirmative)",
        "Affirmative comment. Same bug pattern as C#1, applied to MongoDB aggregation clause. Grade A.",
        None,
    ),
    (
        "ApplicantDetailsPipeline.js", "C#3", 516, "Logic Bug", "major",
        "Affirms removal of parsedSalaryMax > 0 from $lte clause is a correct fix that allows zero as maximum salary.",
        "Ensure try/catch and undefined handling for API-related code",
        "Raised - Correctly (TP, Affirmative)",
        "Affirmative comment. Same bug pattern as C#1/C#2, applied to maximum salary clause. Grade A.",
        None,
    ),
    ("ApplicantDetailsPipeline.js", "-", "-", "-", "-",
     "Flag any console.log, debugger, or unused imports/variables",
     "Flag any console.log, debugger, or unused imports/variables",
     "Not Raised - Justified",
     "No console.log/debugger or unused imports in the diff. Backend file - correctly skipped.",
     "Nothing in the diff to flag"),
    ("ApplicantDetailsPipeline.js", "-", "-", "-", "-",
     "Ensure data-testid is added to new clickable elements",
     "Ensure data-testid is added to new clickable elements",
     "N/A",
     "Backend file (MongoDB pipeline construction) - no clickable UI elements.",
     "Not applicable to this file"),
    ("ApplicantDetailsPipeline.js", "-", "-", "-", "-",
     "Scrutinize useEffect/useCallback dependency arrays",
     "Scrutinize useEffect/useCallback dependency arrays",
     "N/A", "Backend file - no React hooks.", "Not applicable to this file"),
    ("ApplicantDetailsPipeline.js", "-", "-", "-", "-",
     "Suggest next/dynamic for heavy components",
     "Suggest next/dynamic for heavy components",
     "N/A", "Backend file - no components.", "Not applicable to this file"),
    ("ApplicantDetailsPipeline.js", "-", "-", "-", "-",
     "Helpful, senior, concise. Focus on diff. Nudge not block.",
     "Helpful, senior, concise. Focus on diff. Nudge not block.",
     "Followed",
     "All 3 comments concise, focused on the diff, and use affirmative tone. No tone violations.",
     "Followed correctly"),
    (None, None, None, None, None, None, None, None, None, None),

    # ApplicantFilters.js
    ("FILE: ApplicantFilters.js (client/src/components/JobComponent/JobApplicant)", None, None, None, None, None, None, None, None, None),
    (
        "ApplicantFilters.js", "C#1", 27, "Logic Bug", "major",
        "Praises updated handleSliderChange/handleSliderCommit for correctly preventing slider handles from crossing via MIN_DISTANCE clamping.",
        "Helpful, senior, concise. Focus on diff. Nudge not block.",
        "Raised - False Positive (Praise-as-Review)",
        "Praise comment, not a real issue. Also used DIFF POS 27 instead of NEW LINE 105. Grade F.",
        None,
    ),
    (
        "ApplicantFilters.js", "C#2", 15, "Code Quality", "minor",
        "Praises removal of noticePeriods prop and adoption of noticePeriodOptions constant for consistency.",
        "Helpful, senior, concise. Focus on diff. Nudge not block.",
        "Raised - False Positive (Praise-as-Review)",
        "Praise comment, not an issue. Also used DIFF POS 15 instead of NEW LINE 40. Grade F.",
        None,
    ),
    (
        "ApplicantFilters.js", "C#3", 249, "Code Quality", "caution",
        "minWidth increased to 180, maxWidth removed - flag as UI layout change requiring visual verification across screen sizes.",
        "Frontend: Verify SEO and Responsive layout patterns",
        "Raised - Correctly (TP)",
        "Valid caution. Removing maxWidth could cause unexpected expansion. Grade B - could be more specific about what to verify.",
        None,
    ),
    (
        "ApplicantFilters.js", "C#4", 266, "Code Quality", "caution",
        "New fontSize (0.65rem) and fontWeight (500) on Salary Range Typography - flag for design system alignment.",
        "Use Standard MUI Components and Constants",
        "Raised - Correctly (TP)",
        "Valid caution. Arbitrary 0.65rem may not match design system scale. Grade B.",
        None,
    ),
    ("ApplicantFilters.js", "-", "-", "-", "-",
     "Flag any console.log, debugger, or unused imports/variables",
     "Flag any console.log, debugger, or unused imports/variables",
     "Not Raised - Justified",
     "No console.log/debugger or unused imports in the diff.",
     "Nothing in the diff to flag"),
    ("ApplicantFilters.js", "-", "-", "-", "-",
     "Unsafe array access: (sliderDraft || salaryRange)[0/1] when salaryRange is initialised as []",
     "Ensure try/catch and undefined handling for API-related code",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 1)",
     "Major Runtime Error: parent inits salaryRange=[], child accesses [0]/[1] -> NaN propagates and breaks slider. Should be major severity.",
     "Missed by agent - should have been raised (Runtime Error)"),
    ("ApplicantFilters.js", "-", "-", "-", "-",
     "Fragile reliance on event.target.dataset.index in handleSliderCommit",
     "Ensure try/catch and undefined handling for API-related code",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 2)",
     "MUI Slider does not guarantee dataset.index. Else-branch could always run, leading to incorrect min/max adjustment. Minor severity.",
     "Missed by agent - should have been raised (Logic Bug)"),
    ("ApplicantFilters.js", "-", "-", "-", "-",
     "Removal of {noticePeriods.length > 0 && ...} defensive check",
     "Ensure try/catch and undefined handling for API-related code",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 3)",
     "Filter now always renders, will crash if noticePeriodOptions is undefined. Major severity.",
     "Missed by agent - should have been raised (Runtime Error)"),
    ("ApplicantFilters.js", "-", "-", "-", "-",
     "Use ORIG LINE / NEW LINE values from diff table for line numbers",
     "Use ORIG LINE / NEW LINE values from diff table for line numbers",
     "Not Followed",
     "Comments 1 and 2 used DIFF POS values (27, 15) instead of NEW LINE (105, 40).",
     "Line number mapping rule violated"),
    (None, None, None, None, None, None, None, None, None, None),

    # ApplicantResume.js
    ("FILE: ApplicantResume.js (client/src/components/JobComponent/JobApplicant)", None, None, None, None, None, None, None, None, None),
    (
        "ApplicantResume.js", "C#1", 257, "Exception Handling", "minor",
        "Empty catch block on line 257 silently swallows PDF document loading errors; similar empty catches on lines 268 and 275.",
        "Ensure try/catch and undefined handling for API-related code",
        "Raised - Correctly (TP)",
        "Valid finding (catch {} pattern). Effective fix adds console.error logging. Grade A.",
        None,
    ),
    (
        "ApplicantResume.js", "C#2", 378, "Code Quality", "minor",
        "Use of GlobalStyles for component-specific .textLayer can lead to style leakage; suggest sx prop on Box/styled-component.",
        "Use Standard MUI Components and Constants",
        "Raised - Partial TP",
        "Directionally correct (preference for scoped styles), but PDF.js requires the .textLayer global class. Suggested fix is INVALID - applies sx prop to native <div>. Grade C.",
        None,
    ),
    (
        "ApplicantResume.js", "C#3", 164, "Configuration Issue", "minor",
        "Magic number BASE_SCALE=1.4 and zoom limits 0.5/5 should be extracted into named constants.",
        "Use Standard MUI Components and Constants",
        "Raised - Correctly (TP)",
        "Valid finding. Effective fix introduces MIN_ZOOM_SCALE / MAX_ZOOM_SCALE / DEFAULT_BASE_SCALE constants. Grade A.",
        None,
    ),
    ("ApplicantResume.js", "-", "-", "-", "-",
     "console.error(err) on line 333",
     "Flag any console.log, debugger, or unused imports/variables",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 1)",
     "Custom-instruction violation. Agent caught empty catch blocks in same function but missed console.error. caution severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("ApplicantResume.js", "-", "-", "-", "-",
     "Verify new GlobalStyles import alignment with project MUI patterns",
     "Use Standard MUI Components and Constants",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 2)",
     "New import not verified for necessity / consistency with project's MUI conventions. caution severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("ApplicantResume.js", "-", "-", "-", "-",
     "Scrutinize useEffect/useCallback dependency arrays",
     "Scrutinize useEffect/useCallback dependency arrays",
     "Not Raised - Justified",
     "Recommendation 2 mentions verifying renderPages useCallback deps - generic but not a real issue.",
     "No real dependency issues in the diff"),
    ("ApplicantResume.js", "-", "-", "-", "-",
     "If UI changes are detected, remind the author to attach screenshots",
     "Helpful, senior, concise. Focus on diff. Nudge not block.",
     "Raised in Recommendations (Rec #3)",
     "Correctly nudged about screenshots for the new selectable text functionality.",
     "Addressed in recommendations section"),
    (None, None, None, None, None, None, None, None, None, None),

    # ApplicantSummery.js
    ("FILE: ApplicantSummery.js (client/src/components/JobComponent/JobApplicant)", None, None, None, None, None, None, None, None, None),
    (
        "ApplicantSummery.js", "C#1", 175, "Logic Bug", "major",
        "disableHoverListener uses !!!!applicant?.lastAction === seekerApplicationStatus?.APPLICATIONSENT - operator precedence flaw makes the comparison always false.",
        "Ensure try/catch and undefined handling for API-related code",
        "Raised - Correctly (TP)",
        "Excellent catch. Effective fix removes redundant negations. Grade A. Highest-value finding in the file.",
        None,
    ),
    (
        "ApplicantSummery.js", "C#2", 62, "Code Quality", "minor",
        "alignItems changed center -> flex-start and gap=2 added on header Box - asks for visual verification across screen sizes.",
        "Helpful, senior, concise. Focus on diff. Nudge not block.",
        "Raised - False Positive",
        "Deliberate UI layout change, not an issue. Should have been a screenshots nudge in recommendations, not a review comment. Grade F.",
        None,
    ),
    (
        "ApplicantSummery.js", "C#3", 83, "Code Quality", "minor",
        "Name display logic (firstName && firstName) || '' is verbose - suggest nullish coalescing (??).",
        "Helpful, senior, concise. Focus on diff. Nudge not block.",
        "Raised - Partial TP",
        "Redundancy is real but significance is overstated. Effective fix using ??. Grade C.",
        None,
    ),
    ("ApplicantSummery.js", "-", "-", "-", "-",
     "console.error('Error fetching invite links:', error) on line 47",
     "Flag any console.log, debugger, or unused imports/variables",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 1)",
     "Custom-instruction violation. console.error left in production code, should be flagged. minor severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("ApplicantSummery.js", "-", "-", "-", "-",
     "Excessively long nested ternary on line 228 (160+ chars)",
     "Refactor only for high logic density",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 2)",
     "Readability issue, should be extracted into helper / variable. minor severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("ApplicantSummery.js", "-", "-", "-", "-",
     "If UI changes are detected, remind the author to attach screenshots",
     "Helpful, senior, concise. Focus on diff. Nudge not block.",
     "Raised in Recommendations (Rec #1)",
     "Correctly nudged for screenshots across desktop / tablet / mobile views.",
     "Addressed in recommendations section"),
    (None, None, None, None, None, None, None, None, None, None),

    # [jobId].js
    ("FILE: [jobId].js (client/src/pages/recruiter/applicant)", None, None, None, None, None, None, None, None, None),
    (
        "[jobId].js", "C#1", 91, "Integration Issue", "critical",
        "Removing noticePeriods state without updating ApplicantFilters component interface will lead to a runtime error when the component accesses props.noticePeriods.",
        "Ensure try/catch and undefined handling for API-related code",
        "Raised - False Positive (Over-severity, Speculative)",
        "Speculative - cannot see ApplicantFilters internals. Should be 'caution' severity, framed as verification request. Suggested fix is a full revert, not a real fix. Grade F.",
        None,
    ),
    (
        "[jobId].js", "C#2", 88, "Logic Bug", "major",
        "salaryRange initial state changed from [0, 0] to []; fetchApplicants accesses salaryRange[0]/[1] directly, passing undefined to API on first render.",
        "Ensure try/catch and undefined handling for API-related code",
        "Raised - Correctly (TP) but Over-severity",
        "Real issue but should be 'minor' (self-corrects after API). Suggested fix is a revert which may re-introduce the original bug. Grade C.",
        None,
    ),
    ("[jobId].js", "-", "-", "-", "-",
     "console.error('Error fetching applicant details:', error) on line 156",
     "Flag any console.log, debugger, or unused imports/variables",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 1)",
     "Custom-instruction violation - file-wide scan not performed. caution severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("[jobId].js", "-", "-", "-", "-",
     "console.error('Error fetching applicants:', error) on line 173",
     "Flag any console.log, debugger, or unused imports/variables",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 2)",
     "Custom-instruction violation - file-wide scan not performed. caution severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("[jobId].js", "-", "-", "-", "-",
     "console.error('Error fetching cities:', error) on line 196",
     "Flag any console.log, debugger, or unused imports/variables",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 3)",
     "Adjacent to diff changes (same function as line 194). caution severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("[jobId].js", "-", "-", "-", "-",
     "Duplicate API call on mount (salaryRange [] -> [0, fetchedMax])",
     "Scrutinize useEffect/useCallback dependency arrays",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 4)",
     "Performance issue - fetchApplicants fires twice on mount because salaryRange is reset by another useEffect. minor severity.",
     "Missed by agent - should have been raised (Performance Issue)"),
    ("[jobId].js", "-", "-", "-", "-",
     "data-testid audit on modified ApplicantFilters usage",
     "Ensure data-testid is added to new clickable elements",
     "Not Raised - SHOULD HAVE BEEN RAISED (Missed Issue 5)",
     "Component is being modified (prop list changed); reminder to verify data-testid on internal interactive elements is appropriate. caution severity.",
     "Missed by agent - should have been raised (Code Quality)"),
    ("[jobId].js", "-", "-", "-", "-",
     "If UI changes are detected, remind the author to attach screenshots",
     "Helpful, senior, concise. Focus on diff. Nudge not block.",
     "Raised in Recommendations (Rec #2)",
     "Correctly nudged for UI screenshots in PR description.",
     "Addressed in recommendations section"),
]


PER_FILE_STATS = [
    ("ApplicantDetailsPipeline.js", 3, 3, 0, 0, 0,
     "All Logic Bug affirmations (try/catch undefined handling)",
     "None",
     "data-testid, useEffect, next/dynamic, console.log, MUI, Images, SEO",
     "100%", "A"),
    ("ApplicantFilters.js", 4, 2, 2, 0, 3,
     "Responsive layout caution (2), MUI Constants caution",
     "try/catch & undefined handling (3 missed runtime errors), line-number mapping (DIFF POS used)",
     "console.log, useEffect (correct), next/dynamic, secrets, env vars",
     "50%", "D"),
    ("ApplicantResume.js", 3, 2, 0, 1, 2,
     "Empty catch detection, magic-number extraction (Configuration Issue), screenshots nudge",
     "console.error scan (line 333), MUI import verification (GlobalStyles), invalid sx-on-div fix",
     "data-testid, next/dynamic, secrets, env vars, useEffect deps",
     "75%", "B"),
    ("ApplicantSummery.js", 3, 1, 1, 1, 2,
     "Operator-precedence logic bug detection, screenshots nudge",
     "console.error scan (line 47), high-density ternary (line 228), tone (FP UI nudge)",
     "data-testid (no new), useEffect, next/dynamic, secrets, env vars",
     "60%", "B"),
    ("[jobId].js", 2, 1, 1, 0, 5,
     "Integration risk identification, screenshots nudge",
     "Severity decision tree (0% sev acc), file-wide console.* scan (3 missed), useEffect dep duplicate API call, data-testid audit, ineffective revert-style fixes",
     "next/dynamic, secrets, env vars, MUI Components",
     "20%", "D"),
]


OVERALL_SUMMARY = [
    ("Metric", "Value"),
    ("COMMENTS ANALYSIS", None),
    ("Total Comments Raised Across All Files", "15"),
    ("True Positives (correctly raised)", "9 (60.0%)"),
    ("False Positives (incorrectly raised)", "4 (26.7%)"),
    ("Partial True Positives", "2 (13.3%)"),
    ("Severity Accuracy", "73% (11/15)"),
    ("Category Accuracy", "87% (13/15)"),
    ("Line Number Accuracy", "93% (13/15) - 2 used DIFF POS instead of NEW LINE"),
    ("Fix Effectiveness Rate", "44% (effective + partial / total fixes)"),
    ("Total Obvious Misses", "12 across all files"),
    ("Avg Tokens / Valid Insight", "16,640 (very high - low efficiency)"),
    (None, None),
    ("FILE GRADE DISTRIBUTION", None),
    ("Grade A", "1 file (ApplicantDetailsPipeline.js)"),
    ("Grade B", "2 files (ApplicantResume.js, ApplicantSummery.js)"),
    ("Grade C", "0 files"),
    ("Grade D", "2 files (ApplicantFilters.js, [jobId].js)"),
    (None, None),
    ("CUSTOM INSTRUCTIONS COMPLIANCE", None),
    ("Total Custom Instruction Rules Surveyed", "13"),
    (None, None),
    ("Correctly Handled (Followed + N/A)", None),
    ("Fully Followed (raised correct comments)", "4 (try/catch undefined for ApplicantDetailsPipeline + Resume empty catch, MUI constants for ApplicantResume, helpful/concise tone, screenshots reminder in 3 files)"),
    ("N/A - Correctly Not Raised (nothing in diff)", "5 (data-testid backend, next/dynamic backend, console.log clean, secrets, env vars)"),
    ("Total Correct Decisions", "9 / 13 = 69.2%"),
    (None, None),
    ("Partially Correct", None),
    ("Partially Followed (raised but with issues)", "2"),
    ("  - try/catch / undefined handling: caught some, missed 4 critical instances", "ApplicantFilters (3 missed), [jobId].js (1 over-severity revert)"),
    ("  - Tone: nudge not block - violated by FP praise/UI comments", "ApplicantFilters C#1/C#2, ApplicantSummery C#2, [jobId].js C#1 (over-severity)"),
    (None, None),
    ("Failed", None),
    ("Not Followed - Genuine Failure", "2"),
    ("  - Flag any console.log/debugger - missed 5 console.error instances", "ApplicantResume L333, ApplicantSummery L47, [jobId].js L156/L173/L196"),
    ("  - Use ORIG/NEW LINE not DIFF POS", "ApplicantFilters C#1 used 27 (should be 105), C#2 used 15 (should be 40)"),
    (None, None),
    ("OVERALL COMPLIANCE RATE", None),
    ("Fully Correct (Followed + N/A) / Total", "9 / 13 = 69.2%"),
    ("Partially Correct / Total", "2 / 13 = 15.4%"),
    ("Failed / Total", "2 / 13 = 15.4%"),
    ("Combined Pass Rate (Fully + Partially + N/A) / Total", "11 / 13 = 84.6%"),
    (None, None),
    ("GENUINE FAILURES (should have been raised but wasn't)", None),
    ("1. console.error file-wide scan", "5 instances missed across 3 files (ApplicantResume L333, ApplicantSummery L47, [jobId].js L156/L173/L196)"),
    ("2. Unsafe array access pattern", "ApplicantFilters - (sliderDraft || salaryRange)[0/1] when salaryRange initialised as [] - will produce NaN on mount"),
    ("3. Removed defensive check", "ApplicantFilters - {noticePeriods.length > 0 && ...} guard removed; will crash if noticePeriodOptions undefined"),
    ("4. Fragile dataset.index reliance", "ApplicantFilters handleSliderCommit - MUI Slider doesn't guarantee event.target.dataset.index"),
    ("5. Duplicate API call on mount", "[jobId].js - fetchApplicants fires twice because salaryRange is initialised [] then reset to [0, fetchedMax]"),
    ("6. data-testid audit reminder on modified component", "[jobId].js - ApplicantFilters prop list changed but no testid audit nudge"),
    ("7. High logic density (long ternary)", "ApplicantSummery line 228 - 160+ char nested ternary should be extracted"),
    ("8. New MUI import verification", "ApplicantResume - GlobalStyles import not verified for necessity / project conventions"),
    (None, None),
    ("ISSUES IN RAISED COMMENTS (raised but with problems)", None),
    ("1. Praise-as-Review-Comment pattern", "ApplicantFilters C#1/C#2 (slider clamping praise, prop removal praise) - violates 'flag issues, not approve correct code'"),
    ("2. Severity over-escalation", "[jobId].js C#1 critical (should be caution - speculative), C#2 major (should be minor - self-correcting), ApplicantSummery C#2 minor (should be FP/recs)"),
    ("3. Line-number mapping errors (DIFF POS vs NEW LINE)", "ApplicantFilters C#1 used 27 instead of 105, C#2 used 15 instead of 40"),
    ("4. Invalid suggested fix", "ApplicantResume C#2 applies MUI sx prop to native <div> - invalid syntax"),
    ("5. Revert-as-fix pattern", "[jobId].js C#1 and C#2 suggested reverts of intentional changes instead of root-cause analysis"),
    (None, None),
    ("STRENGTHS", None),
    ("Logic-bug detection precision", "100% on Logic Bug category (operator precedence, salary filter affirmations, salaryRange init)"),
    ("Exception handling enforcement", "Empty catch block detected (ApplicantResume) with effective fix"),
    ("Configuration Issue (magic numbers)", "BASE_SCALE / zoom limits flagged with effective constant extraction (ApplicantResume)"),
    ("Screenshots / UI verification reminders", "Correctly nudged in 3/5 files (ApplicantResume, ApplicantSummery, [jobId].js)"),
    ("Affirmative bug-fix recognition", "ApplicantDetailsPipeline - all 3 fixes correctly affirmed"),
    (None, None),
    ("WEAKNESSES", None),
    ("Praise-as-Review false positives", "Distinct from affirmative bug-fix recognition - praising deliberate refactor changes that aren't issues"),
    ("Severity decision tree skipped", "Context Check skipped - speculative claims promoted to major/critical instead of caution"),
    ("File-wide custom-instruction scans not performed", "5 console.error instances missed despite explicit rule; suggests Phase 1 (full file review) not executed"),
    ("Revert-style fixes", "Suggested undoing intentional changes rather than guarding/refactoring"),
    ("Invalid syntactic fixes", "MUI sx prop applied to native HTML elements - tech-stack validation missing"),
    (None, None),
    ("PRODUCTION READINESS VERDICT", None),
    ("Overall", "Ready with fixes - core bug detection valuable but requires prompt refinement to eliminate FPs, enforce severity calibration, and ensure file-wide custom-instruction compliance."),
]


def autosize(ws, max_width=60):
    for col in ws.columns:
        length = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                length = max(length, max(len(line) for line in v.split("\n"))) if v else length
            except Exception:
                pass
        ws.column_dimensions[letter].width = min((length + 2) * 1.1, max_width)


def style_header(ws):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def color_status(ws, col_idx, start_row=2):
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    grey = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            v = (cell.value or "")
            if not isinstance(v, str):
                continue
            if "Correctly (TP" in v or v.startswith("Followed"):
                cell.fill = green
            elif "False Positive" in v or "SHOULD HAVE BEEN" in v or v == "Not Followed":
                cell.fill = red
            elif "Partial" in v:
                cell.fill = yellow
            elif v.startswith("Raised in Recommendations") or v.startswith("Not Raised - Justified"):
                cell.fill = blue
            elif v == "N/A":
                cell.fill = grey


def color_grade(ws, col_idx, start_row=2):
    colors = {
        "A": "C6EFCE",
        "B": "DDEBF7",
        "C": "FFEB9C",
        "D": "FFC7CE",
        "F": "FFC7CE",
    }
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v in colors:
                cell.fill = PatternFill(start_color=colors[v], end_color=colors[v], fill_type="solid")
                cell.font = Font(bold=True)


def style_file_separator(ws, file_col=1, start_row=2):
    fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        v = row[file_col - 1].value
        if isinstance(v, str) and v.startswith("FILE:"):
            for cell in row:
                cell.fill = fill
                cell.font = Font(bold=True, color="FFFFFF")


def main():
    out = Path("c:/Users/USER/pr-eval-agent/custom_instructions_compliance_report_v5.xlsx")

    wb = Workbook()

    # Sheet 1: Comment Analysis
    ws = wb.active
    ws.title = "Comment Analysis"
    headers = ["File", "Comment #", "Line", "Category", "Severity",
               "Review Comment (Summary)", "Maps to Custom Instruction",
               "Status", "Notes", "Reason Not Raised"]
    ws.append(headers)
    for row in COMMENT_ROWS:
        ws.append(list(row))
    style_header(ws)
    color_status(ws, col_idx=8)
    style_file_separator(ws)
    autosize(ws, max_width=70)
    # Wrap long text columns
    wrap_cols = {6, 7, 9, 10}
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="top", wrap_text=(c in wrap_cols)
            )

    # Sheet 2: Per File Stats
    ws2 = wb.create_sheet("Per File Stats")
    ws2.append([
        "File", "Total Comments", "True Positives", "False Positives", "Partial",
        "Obvious Misses", "Custom Instr. Followed", "Custom Instr. Missed",
        "Custom Instr. N/A", "File Compliance %", "Grade",
    ])
    for row in PER_FILE_STATS:
        ws2.append(list(row))
    style_header(ws2)
    color_grade(ws2, col_idx=11)
    autosize(ws2, max_width=55)
    for r in range(2, ws2.max_row + 1):
        for c in range(1, ws2.max_column + 1):
            ws2.cell(row=r, column=c).alignment = Alignment(
                vertical="top", wrap_text=(c >= 7 and c <= 9)
            )

    # Sheet 3: Overall Summary
    ws3 = wb.create_sheet("Overall Summary")
    for row in OVERALL_SUMMARY:
        ws3.append(list(row))
    style_header(ws3)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for r in range(2, ws3.max_row + 1):
        a = ws3.cell(row=r, column=1).value
        b = ws3.cell(row=r, column=2).value
        if isinstance(a, str) and b is None and not a.startswith("  "):
            ws3.cell(row=r, column=1).fill = section_fill
            ws3.cell(row=r, column=1).font = Font(bold=True)
        ws3.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws3.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    ws3.column_dimensions["A"].width = 65
    ws3.column_dimensions["B"].width = 80

    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
