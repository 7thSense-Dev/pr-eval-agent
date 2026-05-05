"""
Generate Excel sheet mapping eval report comments against custom instructions.
"""

import re
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def parse_eval_report(file_path):
    """Parse eval report markdown and extract key data."""
    content = Path(file_path).read_text(encoding='utf-8')
    
    # Extract filename
    filename_match = re.search(r'Evaluation Report: (.+)', content)
    filename = filename_match.group(1).strip('`') if filename_match else "Unknown"
    
    # Extract file path
    file_path_match = re.search(r'\*\*File\*\*: `([^`]+)`', content)
    source_file = file_path_match.group(1) if file_path_match else filename
    
    # Extract summary grid table
    grid_match = re.search(r'### 1a — Summary Grid\n\n\|(.+?)\|\n\n', content, re.DOTALL)
    comments = []
    
    if grid_match:
        table_rows = grid_match.group(1).split('\n')[1:]  # Skip header
        for row in table_rows:
            if not row.strip():
                continue
            parts = [p.strip() for p in row.split('|')[1:-1]]
            if len(parts) >= 8:
                comments.append({
                    'line': parts[0],
                    'category': parts[1],
                    'severity': parts[2],
                    'tp_fp': parts[3],  # TP/FP/P
                    'sev_verdict': parts[4],
                    'cat_verdict': parts[5],
                    'line_check': parts[6],
                    'fix': parts[7],
                })
    
    # Extract stats
    stats_match = re.search(r'\| Total Comments \| (\d+) \|', content)
    total_comments = int(stats_match.group(1)) if stats_match else len(comments)
    
    tp_match = re.search(r'\| True Positives \| (\d+)', content)
    true_positives = int(tp_match.group(1)) if tp_match else 0
    
    fp_match = re.search(r'\| False Positives \| (\d+)', content)
    false_positives = int(fp_match.group(1)) if fp_match else 0
    
    precision_match = re.search(r'\| Precision \| (\d+)% \|', content)
    precision = int(precision_match.group(1)) if precision_match else 0
    
    # Extract grade
    grade_match = re.search(r'\*\*Grade\*\*: ([A-F])', content)
    grade = grade_match.group(1) if grade_match else "N/A"
    
    return {
        'filename': filename,
        'source_file': source_file,
        'comments': comments,
        'total_comments': total_comments,
        'true_positives': true_positives,
        'false_positives': false_positives,
        'precision': precision,
        'grade': grade,
    }


def map_to_custom_instructions():
    """Map categories to custom instruction groupings."""
    return {
        'Code Quality': {
            'group': 'Code Cleanliness & Quality',
            'instructions': [
                'Flag console.log, debugger, unused imports/variables',
                'Ensure data-testid on new clickable elements',
                'Use Standard MUI Components and Constants',
                'Refactor only for high logic density',
            ],
        },
        'Performance & Logic': {
            'group': 'Performance & Logic',
            'instructions': [
                'Scrutinize useEffect/useCallback dependency arrays',
                'Flag heavy computation, missing memoization, debounce/throttle',
                'Ensure try/catch and undefined handling for API calls',
                'Use Image component instead of img in Next.js',
                'Use next/dynamic for heavy components',
            ],
        },
        'Logic Bug': {
            'group': 'Security & Logic',
            'instructions': [
                'Flag logic errors and inconsistencies',
                'Check for race conditions and error handling',
            ],
        },
        'Security': {
            'group': 'Security & Config',
            'instructions': [
                'Flag secrets, API keys, hardcoded credentials',
                'Verify environment variables in README',
                'Flag unnecessary heavy dependencies',
            ],
        },
        'Standards': {
            'group': 'Standards & Consistency',
            'instructions': [
                'Backend: Group responses {success, data, message}',
                'Verify SEO and Responsive layouts',
            ],
        },
    }


def create_excel_sheet(output_dir, reports_data):
    """Create Excel sheet with eval data mapped to custom instructions."""
    output_path = Path(output_dir) / 'eval_report_sheet.xlsx'
    
    # Summary data
    summary_rows = []
    detailed_rows = []
    
    custom_map = map_to_custom_instructions()
    
    for report in reports_data:
        filename = report['filename']
        
        # Add to summary
        summary_rows.append({
            'File': filename,
            'Source': report['source_file'],
            'Total Comments': report['total_comments'],
            'True Positives': report['true_positives'],
            'False Positives': report['false_positives'],
            'Precision': f"{report['precision']}%",
            'Grade': report['grade'],
        })
        
        # Add detailed rows
        for i, comment in enumerate(report['comments'], 1):
            category = comment['category'].strip()
            custom_group = custom_map.get(category, {}).get('group', category)
            
            detailed_rows.append({
                'File': filename,
                'Comment #': i,
                'Line': comment['line'],
                'Category': category,
                'Custom Group': custom_group,
                'Severity': comment['severity'],
                'TP/FP/P': comment['tp_fp'],
                'Sev Verdict': comment['sev_verdict'],
                'Cat Verdict': comment['cat_verdict'],
                'Line Check': comment['line_check'],
                'Fix': comment['fix'],
            })
    
    # Create DataFrames
    summary_df = pd.DataFrame(summary_rows)
    detailed_df = pd.DataFrame(detailed_rows)
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        detailed_df.to_excel(writer, sheet_name='Detailed Comments', index=False)
    
    # Format Excel
    wb = load_workbook(output_path)
    
    # Format Summary sheet
    ws_summary = wb['Summary']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-width columns
    for col in ws_summary.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_summary.column_dimensions[column].width = adjusted_width
    
    # Format Detailed Comments sheet
    ws_detail = wb['Detailed Comments']
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-width columns
    for col in ws_detail.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 50)
        ws_detail.column_dimensions[column].width = adjusted_width
    
    # Color-code TP/FP/P column
    for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row, min_col=7, max_col=7):
        for cell in row:
            if cell.value == 'TP':
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif cell.value == 'FP':
                cell.fill = PatternFill(start_color='FFB6C6', end_color='FFB6C6', fill_type='solid')
            elif cell.value == 'Partial':
                cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
    
    wb.save(output_path)
    return output_path


def main():
    """Main execution."""
    reports_dir = Path('c:/Users/USER/pr-eval-agent/tmp/poc/1627_1777423785050_recrui8/reports_generated')
    
    if not reports_dir.exists():
        print(f"[X] Reports directory not found: {reports_dir}")
        return 1
    
    # Find all eval report markdown files
    report_files = sorted(reports_dir.glob('*_eval_report.md'))
    if not report_files:
        print(f"[X] No eval report files found in {reports_dir}")
        return 1
    
    print(f"\n[*] Found {len(report_files)} eval report file(s)")
    
    # Parse all reports
    reports_data = []
    for report_file in report_files:
        print(f"  [OK] Parsing: {report_file.name}")
        data = parse_eval_report(report_file)
        reports_data.append(data)
    
    # Create Excel sheet
    print(f"\n[*] Generating Excel sheet...")
    output_path = create_excel_sheet(reports_dir, reports_data)
    print(f"[OK] Excel sheet created: {output_path}")
    
    # Print summary stats
    total_all = sum(r['total_comments'] for r in reports_data)
    tp_all = sum(r['true_positives'] for r in reports_data)
    fp_all = sum(r['false_positives'] for r in reports_data)
    
    print(f"\n{'='*60}")
    print("AGGREGATED STATISTICS")
    print(f"{'='*60}")
    print(f"Files Evaluated  : {len(reports_data)}")
    print(f"Total Comments   : {total_all}")
    print(f"True Positives   : {tp_all} ({100*tp_all//max(total_all,1)}%)")
    print(f"False Positives  : {fp_all} ({100*fp_all//max(total_all,1)}%)")
    print(f"Overall Precision: {100*tp_all//max(total_all,1)}%")
    print(f"{'='*60}\n")
    
    return 0


if __name__ == '__main__':
    import sys
    sys.exit(main())
